"""WACC 피어 분석 엑셀 워크북.

원자료만 값(노란 배경), 계산은 전부 수식(남색 글자). 요약!B4(베타 소스)·B5(D/E 방식)
셀만 바꾸면 전체가 재계산된다. 설명 문자열은 절대 '=' 로 시작하지 않는다
(openpyxl이 수식으로 저장해 '복구' 프롬프트를 띄운다).
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import srp as _srp
from kofia import _MAT as KOFIA_MATURITIES

FONT = "맑은 고딕"
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
FORMULA_FONT = Font(name=FONT, size=10, color="203864")
BASE_FONT = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

DEBT_COLS = ["단기차입금", "유동성장기부채", "장기차입금", "사채", "리스부채", "기타"]
PEER_HEADERS = ["회사명", "종목코드", "KSIC", "시장", "종가(KRX)", "종가(한공회)", "종가 적용",
                "보통주 유통주식수", "우선주 유통주식수", "비지배지분", "시가자기자본(E)",
                "단기차입금", "유동성장기부채", "장기차입금", "사채(전환·BW·교환 포함)", "리스부채", "기타 차입",
                "이자부부채(D)", "D/E", "세율", "βL 산출(Blume)", "βL 한공회(조정)", "βL 적용", "βU",
                "집계포함", "플래그", "rcept_no", "보고서", "연결/별도", "세전이익(과표대용)", "세율 근거", "감사의견", "영업이익(최근 FY)"]
BETA_HEADERS = ["회사명", "종목코드", "raw β", "R²", "n", "Blume β", "관측치",
                "한공회 기준일자", "한공회 종가", "한공회 실질베타", "한공회 조정베타", "한공회 포인트수", "차이(한공회−산출)"]


class _W:
    """워크북 쓰기 헬퍼: 값/수식/헤더를 구분해 서식과 수식 좌표 집합을 관리한다."""

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.formulas = set()

    def sheet(self, name):
        ws = self.wb.create_sheet(name)
        ws.sheet_view.showGridLines = True
        return ws

    def header(self, ws, row, values):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.font, cell.fill, cell.border = BOLD, HEADER_FILL, BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def val(self, ws, row, col, v, *, inp=False, fmt=None):
        cell = ws.cell(row, col, v)
        cell.font, cell.border = BASE_FONT, BORDER
        if inp:
            cell.fill = INPUT_FILL
        if fmt:
            cell.number_format = fmt
        return cell

    def f(self, ws, row, col, formula, fmt=None):
        assert formula.startswith("=")
        cell = ws.cell(row, col, formula)
        cell.font, cell.border = FORMULA_FONT, BORDER
        if fmt:
            cell.number_format = fmt
        self.formulas.add(f"{ws.title}!{cell.coordinate}")
        return cell

    def text(self, ws, row, col, s):
        assert not str(s).startswith("=")
        cell = ws.cell(row, col, s)
        cell.font = BASE_FONT
        return cell

    def widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _summary(w, d):
    ws = w.sheet("요약")
    ws["A1"] = f"WACC 피어 분석 — {d['target']['name']}"
    ws["A1"].font = Font(name=FONT, size=13, bold=True)
    rows = [("대상기업", d["target"]["name"]), ("기준일", d["as_of"]),
            ("베타 소스", "한공회" if d["beta_source"] == "kicpa" else "산출"),
            ("목표 D/E 방식", "중앙값" if d["de_method"] == "median" else "평균"),
            ("대상 세율", d["tax_target"]),
            ("리스부채 포함(D)", bool(d.get("include_lease", True)))]
    for i, (k, v) in enumerate(rows, 2):
        w.text(ws, i, 1, k)
        w.val(ws, i, 2, v, inp=(i >= 4), fmt="0.0%" if i == 6 else None)
    dv1 = DataValidation(type="list", formula1='"한공회,산출"', allow_blank=False)
    dv2 = DataValidation(type="list", formula1='"평균,중앙값"', allow_blank=False)
    dv3 = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv1); ws.add_data_validation(dv2); ws.add_data_validation(dv3)
    dv1.add("B4"); dv2.add("B5"); dv3.add("B7")
    w.text(ws, 8, 1, "βU 평균(집계)")
    w.f(ws, 8, 2, '=IF($B$5="중앙값",MEDIAN(목표DE!$C:$C),AVERAGE(목표DE!$C:$C))', "0.0000")
    w.text(ws, 9, 1, "목표 D/E")
    w.f(ws, 9, 2, '=IF($B$5="중앙값",MEDIAN(목표DE!$D:$D),AVERAGE(목표DE!$D:$D))', "0.0000")
    w.text(ws, 10, 1, "재레버 βL")
    w.f(ws, 10, 2, "=B8*(1+(1-B6)*B9)", "0.0000")
    w.text(ws, 11, 1, "집계 피어 수")
    w.f(ws, 11, 2, "=COUNT(목표DE!$C:$C)")
    notes = [
        "규약: 2년 주간(104주, 기준일 직전 금요일 마감) 수익률을 KOSPI에 회귀 → Blume(2/3·raw+1/3) → Hamada 언레버 βU=βL/(1+(1−t)·D/E) → 피어 집계 → 목표 D/E로 재레버.",
        "E(시가자기자본) = 종가 × (보통주 유통주식수 + 우선주 유통주식수) + 비지배지분 장부가. D = '이자부부채 산정내역' 시트의 '이자부부채 합계'(F열 분류가 원천: 규칙 판정 + 사용자 조정). 리스부채 포함 여부는 B7.",
        f"조회일 {d.get('query_date', '')} · 주가 구간 {d.get('window', ('', ''))[0]}~{d.get('window', ('', ''))[1]} · 출처: KRX 정보데이터시스템(주가·시총), DART OpenAPI(재무·주식수·KSIC), 한공회 베타 조회 파일.",
        "세율: 피어별로 최신 사업보고서 법인세비용차감전순이익을 과세표준 대용으로 2026년 구간표(11.0/22.0/24.2/27.5%, 지방세 포함)에 대입한 한계세율. 결손·미확인은 대상 세율 적용(피어 시트 '세율 근거' 열). 피어 시트 세율 셀은 직접 수정 가능.",
        "한계: 주가는 배당락 미보정 단순 종가 수익률. 우선주 시세는 보통주 종가 준용.",
        f"키워드: {', '.join(d.get('keywords', []))}",
        "WACC 시트: Rf·Kd는 '금투협' 시트 원자료(노란 셀)를 수식으로 참조해 자동 채움(입력 셀이라 수정 가능). "
        "SRP는 대상 시총(후보군 탭)·순자산(요약 상단, 억)으로 한공회 연구결과(260605) 5분위를 자동 판정.",
    ] + list(d.get("notes", []))
    for i, n in enumerate(notes, 13):
        w.text(ws, i, 1, n)
    w.widths(ws, [22, 60])
    ws.freeze_panes = "A2"


def _candidates(w, d):
    ws = w.sheet("후보군")
    kws = list(d.get("keywords", []))
    k0 = 6                                   # F
    excl = k0 + len(kws)                     # 제외
    reason, target, listed, settle, flags, hits = excl + 1, excl + 2, excl + 3, excl + 4, excl + 5, excl + 6
    w.header(ws, 1, ["회사명", "시장", "종목코드", "업종", "주요제품"] + kws
             + ["제외", "제외 사유", "대상", "상장일", "결산월", "플래그", "적중"])
    for r, c in enumerate(d.get("candidates", []), 2):
        for col, key in enumerate(["name", "market", "code", "industry", "products"], 1):
            w.val(ws, r, col, c.get(key, ""))
        for j in range(len(kws)):
            L = get_column_letter(k0 + j)
            w.f(ws, r, k0 + j, f'=IFERROR(FIND({L}$1,$E{r}&" "&$D{r}),0)>0')
        w.val(ws, r, excl, True if c.get("excluded") else None, inp=True)
        w.val(ws, r, reason, c.get("reason", ""), inp=True)
        if kws:
            w.f(ws, r, target, f"=AND(OR({get_column_letter(k0)}{r}:{get_column_letter(excl - 1)}{r}),NOT({get_column_letter(excl)}{r}=TRUE))")
        else:
            w.f(ws, r, target, f"=NOT({get_column_letter(excl)}{r}=TRUE)")
        w.val(ws, r, listed, c.get("listed", ""))
        w.val(ws, r, settle, c.get("settle_month", ""))
        w.val(ws, r, flags, (("★ " if c.get("recommended") else "") + c["reason"]) if c.get("reason") else ", ".join(c.get("flags", [])))
        extra = ([f"KSIC{'≡' if c.get('ksic_level') == 2 else '≈'}"] if c.get("ksic_level") else []) + (["시총유사"] if c.get("cap_similar") else []) \
                + [f"제외어:{k}" for k in (c.get("neg_hits") or [])] + (["★추천"] if c.get("recommended") else [])
        w.val(ws, r, hits, ", ".join(list(c.get("hits", [])) + extra))
    n = len(d.get("candidates", []))
    ws.auto_filter.ref = f"A1:{get_column_letter(hits)}{max(n + 1, 2)}"
    ws.freeze_panes = "D2"
    w.widths(ws, [18, 7, 9, 30, 40] + [8] * len(kws) + [6, 16, 6, 11, 7, 20, 14])


def _prices(w, d, peers):
    ws = w.sheet("주가")
    dates, idx = d["weekly_dates"], d["weekly_index"]
    n = len(dates)
    codes = [p["code"] for p in peers]
    names = [p["name"] for p in peers]
    w.header(ws, 1, ["주차 종료일", "KOSPI"] + names)
    for i in range(n):
        w.val(ws, i + 2, 1, dates[i])
        w.val(ws, i + 2, 2, idx[i], inp=True, fmt="#,##0.00")
        for j, code in enumerate(codes):
            v = d["weekly"].get(code, [None] * n)[i]
            w.val(ws, i + 2, 3 + j, v, inp=True, fmt="#,##0")
    gap = 3 + len(codes)                          # 빈 열
    r0 = gap + 1                                  # 수익률 시작 열
    w.header(ws, 1, [None] * (gap - 1) + [None] + ["KOSPI 수익률"] + [f"{nm} 수익률" for nm in names])
    ws.cell(1, gap).value = None
    for i in range(1, n):
        row = i + 2
        for j in range(len(codes) + 1):
            src = get_column_letter(2 + j)
            w.f(ws, row, r0 + j, f'=IF(OR({src}{row - 1}="",{src}{row}=""),"",{src}{row}/{src}{row - 1}-1)', "0.00%")
    ws.freeze_panes = "B2"
    ranges = {}
    last = n + 1
    idx_rng = f"${get_column_letter(r0)}$3:${get_column_letter(r0)}${last}"
    for j, code in enumerate(codes):
        L = get_column_letter(r0 + 1 + j)
        ranges[code] = (f"${L}$3:${L}${last}", idx_rng)
    return ranges


def _beta(w, d, peers, ranges):
    ws = w.sheet("베타")
    w.header(ws, 1, BETA_HEADERS)
    for r, p in enumerate(peers, 2):
        stk, idx = ranges[p["code"]]
        w.val(ws, r, 1, p["name"]); w.val(ws, r, 2, p["code"])
        w.f(ws, r, 3, f"=SLOPE(주가!{stk},주가!{idx})", "0.0000")
        w.f(ws, r, 4, f"=RSQ(주가!{stk},주가!{idx})", "0.0000")
        w.f(ws, r, 5, f"=COUNT(주가!{stk})")
        w.f(ws, r, 6, f'=IF(C{r}="","",2/3*C{r}+1/3)', "0.0000")
        w.f(ws, r, 7, f'=IF(E{r}<104,"관측치부족","")')
        k = p.get("kicpa") or {}
        w.val(ws, r, 8, k.get("base_date"), inp=True)
        w.val(ws, r, 9, k.get("close"), inp=True, fmt="#,##0")
        w.val(ws, r, 10, k.get("raw"), inp=True, fmt="0.000000")
        w.val(ws, r, 11, k.get("adjusted"), inp=True, fmt="0.000000")
        w.val(ws, r, 12, k.get("points"), inp=True)
        w.f(ws, r, 13, f'=IF(OR(K{r}="",F{r}=""),"",K{r}-F{r})', "0.0000")
    ws.freeze_panes = "C2"
    w.widths(ws, [18, 9] + [10] * 11)


def _peers(w, d, peers, total_rows=None):
    ws = w.sheet("피어")
    total_rows = total_rows or {}
    w.header(ws, 1, PEER_HEADERS)
    for r, p in enumerate(peers, 2):
        w.val(ws, r, 1, p["name"]); w.val(ws, r, 2, p["code"]); w.val(ws, r, 3, p.get("ksic", "")); w.val(ws, r, 4, p.get("market", ""))
        w.val(ws, r, 5, p.get("close_krx"), inp=True, fmt="#,##0")
        w.val(ws, r, 6, p.get("close_kicpa"), inp=True, fmt="#,##0")
        w.f(ws, r, 7, f'=IF(AND(요약!$B$4="한공회",F{r}<>""),F{r},E{r})', "#,##0")
        w.val(ws, r, 8, p.get("common_out", 0), inp=True, fmt="#,##0")
        w.val(ws, r, 9, p.get("pref_out", 0), inp=True, fmt="#,##0")
        w.val(ws, r, 10, p.get("nci", 0), inp=True, fmt="#,##0")
        w.f(ws, r, 11, f"=G{r}*(H{r}+I{r})+J{r}", "#,##0")
        for j, key in enumerate(DEBT_COLS):
            cls = "리스부채" if key == "리스부채" else "이자부부채"
            w.f(ws, r, 12 + j, f'=SUMIFS({DEBT_SHEET_REF}!$E:$E,{DEBT_SHEET_REF}!$H:$H,$B{r},{DEBT_SHEET_REF}!$I:$I,"{key}",{DEBT_SHEET_REF}!$F:$F,"{cls}")*1000000', "#,##0")
        tr = total_rows.get(p["code"])
        w.f(ws, r, 18, f"={DEBT_SHEET_REF}!E{tr}*1000000" if tr else f"=SUM(L{r}:O{r})+Q{r}+IF(요약!$B$7=TRUE,P{r},0)", "#,##0")
        w.f(ws, r, 19, f'=IF(K{r}=0,"",R{r}/K{r})', "0.0000")
        w.val(ws, r, 20, p.get("tax", d["tax_target"]), inp=True, fmt="0.0%")
        w.f(ws, r, 21, f"=베타!F{r}", "0.0000")
        w.f(ws, r, 22, f"=베타!K{r}", "0.0000")
        w.f(ws, r, 23, f'=IF(AND(요약!$B$4="한공회",V{r}<>""),V{r},U{r})', "0.0000")
        w.f(ws, r, 24, f'=IF(OR(W{r}="",S{r}=""),"",W{r}/(1+(1-T{r})*S{r}))', "0.0000")
        w.val(ws, r, 25, bool(p.get("include")), inp=True)
        w.val(ws, r, 26, ", ".join(p.get("flags", [])))
        w.val(ws, r, 27, p.get("rcept_no")); w.val(ws, r, 28, p.get("report_label", "")); w.val(ws, r, 29, p.get("fs_div"))
        w.val(ws, r, 30, p.get("pretax"), fmt="#,##0"); w.val(ws, r, 31, p.get("tax_basis", ""))
        w.val(ws, r, 32, p.get("audit_opinion")); w.val(ws, r, 33, p.get("op_income"), fmt="#,##0")
    ws.freeze_panes = "C2"
    w.widths(ws, [18, 9, 7, 6, 10, 10, 10, 14, 12, 12, 16] + [12] * 6 + [14, 8, 7, 10, 10, 10, 10, 8, 20, 16, 14, 8, 16, 30, 12, 16])


# ── WACC 시트: 기존 실무 조서 양식 복제 ─────────
# 맑은 고딕 9, 격자선 끔, 줌 85. 1행 검정 제목 바, 좌측 피어 표(D:I, 입력열 노랑, βU 열 굵은 테두리),
# 좌측 요약 상자(D:E, 법인세율·Rf 노랑, Rm·Specific Risk 연두, 평균βU·WACC 회색), 우측 파란 헤더 블록(M:W).
# 피어 5개면 원본과 셀 좌표가 동일하고, n개면 아래로 늘어난다.
W9 = Font(name=FONT, size=9)
W9B = Font(name=FONT, size=9, bold=True)
W9I = Font(name=FONT, size=9, italic=True)
W9_TITLE = Font(name=FONT, size=9, bold=True, color="FFFFFF")
W9_HDR = Font(name=FONT, size=9, bold=True, color="002060")
FILL_NAVY = PatternFill("solid", fgColor="1F3864")       # 제목 바
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")      # 입력 셀(연노랑)
FILL_SOFT = PatternFill("solid", fgColor="EAF1FA")       # 결과 강조(옅은 하늘)
FILL_HDR = PatternFill("solid", fgColor="C6D9F1")        # 헤더 하늘색
THIN_HAIR = Side(style="hair", color="9AA8BF")
THIN_S = Side(style="thin"); MED_S = Side(style="medium")
WACC_WIDTHS = {"A": 2.0, "B": 13.0, "C": 13.0, "D": 24.5, "E": 14.75, "F": 13.0, "G": 13.0, "H": 13.0, "I": 13.0,
               "J": 11.12, "K": 13.0, "L": 13.0, "M": 21.0, "N": 9.62, "O": 11.0, "P": 13.0, "Q": 6.62, "R": 7.5,
               "S": 6.62, "T": 13.0, "U": 13.0, "V": 13.0, "W": 13.0}
VC = Alignment(vertical="center")
CC = Alignment(horizontal="center", vertical="center")
RC = Alignment(horizontal="right", vertical="center")
LC = Alignment(horizontal="left", vertical="center")


def _wacc(w, d, peers, kofia_refs=None):
    """WACC 시트 — 한 흐름 레이아웃. ① 유사기업 베타 표(한 번만) → ② 자본비용 산출(항목·값·근거) + 수식 스트립 + WACC 콜아웃.
    스타일: 남색 제목 바, 하늘색 헤더, 헤어라인, 입력 셀만 연노랑. 좌측 표는 피어·요약 시트를 직접 참조하고 ②는 ①을 참조한다.
    kofia_refs 가 있으면 Rf·Kd 입력 셀이 '금투협' 시트 원자료 셀을 수식으로 참조한다(없으면 값 채움)."""
    ws = w.sheet("WACC")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    for col, wd in {"A": 2, "B": 26, "C": 13, "D": 34, "E": 14, "F": 13, "G": 4.5, "H": 13, "I": 4.5, "J": 13,
                    "K": 4.5, "L": 13, "M": 4.5, "N": 13}.items():
        ws.column_dimensions[col].width = wd
    NAVY, HDR_T = "1F3864", "002060"
    F9, F9B = Font(name=FONT, size=9), Font(name=FONT, size=9, bold=True)
    F9N = Font(name=FONT, size=9, color="6B7A90")
    F_HDR = Font(name=FONT, size=9, bold=True, color=HDR_T)
    F_SEC = Font(name=FONT, size=10, bold=True, color=NAVY)
    F_TITLE = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    HAIR = Side(style="hair", color="9AA8BF"); THIN = Side(style="thin", color="1F3864"); MED = Side(style="medium", color="1F3864")
    VC, CC, RC = Alignment(vertical="center"), Alignment(horizontal="center", vertical="center", wrap_text=True), Alignment(horizontal="right", vertical="center")

    def cell(ref, v=None, *, font=F9, fmt=None, fill=None, align=VC, border=None):
        c = ws[ref]
        if v is not None:
            c.value = v
            if isinstance(v, str) and v.startswith("="):
                c.data_type = "s"
        c.font, c.alignment = font, align
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        if border: c.border = border
        return c

    def f(ref, formula, *, fmt=None, font=F9, fill=None, align=RC, border=None):
        assert formula.startswith("=")
        c = ws[ref]; c.value = formula; c.font, c.alignment = font, align
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        if border: c.border = border
        w.formulas.add(f"WACC!{ref}")
        return c

    def hdr(ref, v):
        return cell(ref, v, font=F_HDR, fill=FILL_HDR, align=CC, border=Border(bottom=MED))

    def section(row, text):
        for col in "BCDEFGHIJKLMN":
            cell(f"{col}{row}", border=Border(bottom=THIN))
        cell(f"B{row}", text, font=F_SEC, border=Border(bottom=THIN))
        ws.row_dimensions[row].height = 20

    n = max(len(peers), 1)
    # ── 제목 바 ──
    ws.row_dimensions[1].height = 26
    for col in "ABCDEFGHIJKLMN":
        cell(f"{col}1", font=F_TITLE, fill=FILL_NAVY)
    ws["A1"].value = f"  WACC 산출 — {d['target']['name']} · 기준일 {d['as_of']}"
    cell("B2", "2년 주간 β(Blume 조정) → Hamada 언레버 → 피어 평균 βU → 목표 D/E로 재레버 · 세율은 피어별 한계세율 · 출처: KRX 시세, DART 재무, 한공회 가이던스", font=F9N)

    # ── ① 유사기업 베타 ──
    section(4, "①  유사기업 베타 (Peer Beta)")
    ws.row_dimensions[5].height = 30
    for ref, v in (("B5", "회사명"), ("C5", "종목코드"), ("D5", "대용기업"), ("E5", "Beta L\n(2y weekly)"), ("F5", "시가총액\n(백만원)"),
                   ("G5", ""), ("H5", "이자부부채\n(백만원)"), ("I5", ""), ("J5", "D/E"), ("K5", ""), ("L5", "세율"), ("M5", ""), ("N5", "Beta U")):
        hdr(ref, v)
    ws.merge_cells("F5:G5"); ws.merge_cells("H5:I5")
    r0 = 6
    for i, p in enumerate(peers):
        r, pr = r0 + i, 2 + i
        last = i == n - 1
        bd = Border(bottom=THIN if last else HAIR)
        cell(f"B{r}", p["name"], font=F9B, border=bd)
        cell(f"C{r}", p["code"], border=bd, align=CC)
        f(f"D{r}", f'=IF(피어!Y{pr}=TRUE,"O","")', align=CC, border=bd)
        f(f"E{r}", f"=피어!W{pr}", fmt="0.0000", border=bd)
        f(f"F{r}", f"=피어!K{pr}/1000000", fmt="#,##0", border=bd); cell(f"G{r}", border=bd)
        f(f"H{r}", f"=피어!R{pr}/1000000", fmt="#,##0", border=bd); cell(f"I{r}", border=bd)
        f(f"J{r}", f'=IF(D{r}="O",피어!S{pr},"")', fmt="0.00%", border=bd); cell(f"K{r}", border=bd)
        f(f"L{r}", f"=피어!T{pr}", fmt="0.0%", border=bd); cell(f"M{r}", border=bd)
        f(f"N{r}", f'=IFERROR(E{r}/(1+(1-L{r})*J{r}),"")', fmt="0.0000", font=F9B, border=bd)
    ra = r0 + n                                            # 평균 행
    lo, hi = r0, ra - 1
    for col in "BCDEFGHIJKLMN":
        cell(f"{col}{ra}", fill=FILL_SOFT, border=Border(bottom=THIN))
    cell(f"B{ra}", "평균 (집계 대상 O)", font=F9B, fill=FILL_SOFT, border=Border(bottom=THIN))
    f(f"J{ra}", f'=IF(요약!$B$5="중앙값",MEDIAN(J{lo}:J{hi}),AVERAGE(J{lo}:J{hi}))', fmt="0.00%", font=F9B, fill=FILL_SOFT, border=Border(bottom=THIN))
    f(f"N{ra}", f'=IF(요약!$B$5="중앙값",MEDIAN(N{lo}:N{hi}),AVERAGE(N{lo}:N{hi}))', fmt="0.0000", font=F9B, fill=FILL_SOFT, border=Border(bottom=THIN))
    cell(f"C{ra}", '=IF(요약!$B$5="중앙값","중앙값","평균")', fill=FILL_SOFT, border=Border(bottom=THIN))
    w.formulas.add(f"WACC!C{ra}"); ws[f"C{ra}"].data_type = "f"; ws[f"C{ra}"].font = F9N

    # ── ② 자본비용 산출 ──
    s = ra + 2
    section(s, "②  자본비용 산출 (Cost of Capital)")
    h = s + 1
    for ref, v in ((f"B{h}", "항목"), (f"C{h}", "값"), (f"D{h}", "근거 · 입력")):
        hdr(ref, v)
    rates = d.get("rates") or {}
    refs = kofia_refs or {}
    rf_v = rates.get("ktb10_final") or rates.get("ktb10_val")
    kd_v = rates.get("bbb_minus_5y")
    used = rates.get("date_used", "")
    rf_note = (f"금투협 최종호가수익률 국고채 10년 {rf_v:.3f}% ({used})" if rf_v else "국고채 10년 — 금투협/ECOS 기준일 값 입력")
    kd_note = (f"금투협 시가평가(평가사 평균) 회사채 무보증 BBB- 5년 {kd_v:.3f}% ({used})" if kd_v
               else "무보증 회사채 BBB- 5년 — 금투협 시가평가 입력")
    rf_val = f"=금투협!{refs['rf']}/100" if refs.get("rf") else (rf_v / 100 if rf_v else None)
    kd_val = f"=금투협!{refs['kd']}/100" if refs.get("kd") else (kd_v / 100 if kd_v else None)
    t_info = d.get("target") or {}
    sj = _srp.judge(cap_million=(t_info.get("cap_eok") or 0) * 100 or None,
                    net_assets_million=(t_info.get("net_assets_eok") or 0) * 100 or None)
    items = [
        ("언레버 베타 (βU, 피어 평균)", f"=N{ra}", "0.0000", None, "① 표 평균", False),
        ("목표부채비율 (D/E)", f"=J{ra}", "0.00%", None, "① 표 평균 (요약!B5 평균/중앙값)", False),
        ("법인세율 (대상회사)", "=요약!$B$6", "0.00%", None, "요약!B6", False),
        ("레버리지 베타 (βL)", None, "0.0000", None, "βU × (1 + (1−t) × D/E)", True),
        ("무위험이자율 (Rf)", rf_val, "0.00%", FILL_INPUT, rf_note, False),
        ("시장위험프리미엄 (MRP)", 0.08, "0.00%", FILL_INPUT, "한공회 가이던스 7~9% 중 8%", False),
        ("기업규모 프리미엄 (SRP)", sj["premium"], "0.00%", FILL_INPUT, sj["note"], False),
        ("자기자본비용 (Ke)", None, "0.00%", None, "Rf + βL × MRP + SRP", True),
        ("타인자본비용 (Kd)", kd_val, "0.000%", FILL_INPUT, kd_note, False),
        ("세후 타인자본비용 Kd(1−t)", None, "0.00%", None, "Kd × (1 − t)", False),
        ("자기자본 비중 E/V", None, "0.00%", None, "1 / (1 + D/E)", False),
        ("타인자본 비중 D/V", None, "0.00%", None, "D/E / (1 + D/E)", False),
        ("WACC", None, "0.00%", FILL_HDR, "Ke × E/V + Kd(1−t) × D/V", True),
    ]
    R = {}
    for k, (label, val, fmt, fill, note, strong) in enumerate(items):
        R[label] = h + 1 + k
    bu, de, t, bl = R["언레버 베타 (βU, 피어 평균)"], R["목표부채비율 (D/E)"], R["법인세율 (대상회사)"], R["레버리지 베타 (βL)"]
    rf, mrp, srp, ke = R["무위험이자율 (Rf)"], R["시장위험프리미엄 (MRP)"], R["기업규모 프리미엄 (SRP)"], R["자기자본비용 (Ke)"]
    kd, kdt, ev, dv, wacc = R["타인자본비용 (Kd)"], R["세후 타인자본비용 Kd(1−t)"], R["자기자본 비중 E/V"], R["타인자본 비중 D/V"], R["WACC"]
    formulas = {bl: f"=C{bu}*(1+(1-C{t})*C{de})", ke: f"=C{rf}+C{bl}*C{mrp}+C{srp}", kdt: f"=C{kd}*(1-C{t})",
                ev: f"=1/(1+C{de})", dv: f"=C{de}/(1+C{de})", wacc: f"=ROUND(C{ke}*C{ev}+C{kdt}*C{dv},4)"}
    for k, (label, val, fmt, fill, note, strong) in enumerate(items):
        r = h + 1 + k
        last = label == "WACC"
        bd = Border(bottom=THIN if last else HAIR, top=THIN if last else None)
        lf = F9B if strong else F9
        cell(f"B{r}", label, font=lf, fill=fill if last else None, border=bd)
        if r in formulas:
            f(f"C{r}", formulas[r], fmt=fmt, font=lf, fill=fill, border=bd)
        elif isinstance(val, str) and val.startswith("="):
            f(f"C{r}", val, fmt=fmt, font=lf, fill=fill, border=bd)
        else:
            cell(f"C{r}", val, fmt=fmt, fill=fill, border=bd, align=RC)
        cell(f"D{r}", note, font=F9N, fill=fill if last else None, border=bd)
    ws[f"C{wacc}"].font = Font(name=FONT, size=10, bold=True, color=NAVY)

    # ── 수식 스트립 (② 오른쪽) ──
    def strip(row, parts, vals):
        cols = "FGHIJKLMN"
        for col, txt in zip(cols, parts):
            hdr(f"{col}{row}", txt)
        for col, v in zip(cols, vals):
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("=") and len(v) > 1:
                f(f"{col}{row+1}", v, fmt="0.00%" if col != "J" or row != h else "0.00", align=CC, border=Border(bottom=HAIR))
            else:
                cell(f"{col}{row+1}", v, align=CC, border=Border(bottom=HAIR))
        ws.row_dimensions[row].height = 30
    strip(h, ["Rf", "+", "MRP", "×", "βL", "+", "SRP", "=", "Ke"],
          [f"=C{rf}", "+", f"=C{mrp}", "×", f"=C{bl}", "+", f"=C{srp}", "=", f"=C{ke}"])
    ws[f"J{h+1}"].number_format = "0.00"; ws[f"N{h+1}"].font = F9B
    strip(h + 4, ["Ke", "×", "E/V", "+", "Kd(1−t)", "×", "D/V", "=", "WACC"],
          [f"=C{ke}", "×", f"=C{ev}", "+", f"=C{kdt}", "×", f"=C{dv}", "=", f"=C{wacc}"])
    ws[f"N{h+5}"].font = F9B
    # ── WACC 콜아웃 ──
    c0 = h + 8
    ws.merge_cells(f"F{c0}:J{c0+1}"); ws.merge_cells(f"K{c0}:N{c0+1}")
    for col in "FGHIJKLMN":
        for rr in (c0, c0 + 1):
            cell(f"{col}{rr}", fill=FILL_SOFT)
    cell(f"F{c0}", "가중평균자본비용 (WACC)", font=Font(name=FONT, size=11, bold=True, color=NAVY), fill=FILL_SOFT, align=CC)
    f(f"K{c0}", f"=C{wacc}", fmt="0.00%", font=Font(name=FONT, size=16, bold=True, color=NAVY), fill=FILL_SOFT, align=CC)
    for col in "FGHIJKLMN":
        ws[f"{col}{c0}"].border = Border(top=MED, left=MED if col == "F" else None, right=MED if col == "N" else None)
        ws[f"{col}{c0+1}"].border = Border(bottom=MED, left=MED if col == "F" else None, right=MED if col == "N" else None)
    ws.row_dimensions[c0].height = 22; ws.row_dimensions[c0 + 1].height = 22
    src = "Source: KRX·한국거래소 시세, DART OpenAPI, 금투협 채권정보센터(최종호가·시가평가), 한공회 가이던스 · 연노랑 셀은 입력값(자동 채움 포함, 수정 가능)"
    if rates:
        src += (f"  |  참고: 시가평가 국고 10년 {rates.get('ktb10_val')}% · BBB- 3년 {rates.get('bbb_minus_3y')}% · AA- 5년 {rates.get('aa_minus_5y')}% ({used})")
    cell(f"B{wacc+2}", src, font=Font(name=FONT, size=8, italic=True, color="6B7A90"))
    ws.freeze_panes = "A3"


DEBT_SHEET = "이자부부채 산정내역"
DEBT_SHEET_REF = f"'{DEBT_SHEET}'"
_CLASSES = ["이자부부채", "리스부채", "제외", "현금및현금성자산", "단기금융상품"]


def _debt_schedule(w, d, peers):
    """조서 '4.이자부부채 산정내역' 양식. 회사 블록 = 계정 행들 + 소계 행. 단위 백만원.
    F열(분류)이 원천 — 바꾸면 소계 → 피어 시트 → WACC 시트가 재계산된다. H·I열은 숨김 도우미(종목코드·버킷).
    반환: {code: 이자부부채 합계 행번호} (피어 시트가 참조)."""
    import pipeline as _pl
    ws = w.sheet(DEBT_SHEET)
    ws.sheet_view.showGridLines = False
    for col, wd in {"A": 3, "B": 20, "C": 15, "D": 34, "E": 16, "F": 14, "G": 46, "H": 10, "I": 12}.items():
        ws.column_dimensions[col].width = wd
    ws.column_dimensions["H"].hidden = True; ws.column_dimensions["I"].hidden = True
    F10, F10B = Font(name=FONT, size=10), Font(name=FONT, size=10, bold=True)
    note = Font(name=FONT, size=9, color="595959"); link = Font(name=FONT, size=10, color="0563C1")
    hdr_fill, sub_fill = PatternFill("solid", fgColor="1F3864"), PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin"); box = Border(top=thin, bottom=thin, left=thin, right=thin)
    AMT = r'#,##0;[Red]\(#,##0\)'
    ws["B1"] = "이자부부채 · 현금성자산 산정 내역 (계정별)"; ws["B1"].font = Font(name=FONT, size=14, bold=True)
    fs_date = d.get("fs_as_of") or d["as_of"]
    ws["B2"] = (f"출처: DART 재무상태표 계정(기간말 ≤ 재무 기준일 {fs_date} 인 최신 보고서, 조회일 {d.get('query_date', '')}). 단위 백만원. "
                "규칙 판정 = 계정명에 '차입/사채/리스부채/유동성장기부채' 포함, '매입채무·미지급·예수·충당' 제외.")
    ws["B2"].font = note
    ws["B3"] = ("F열(분류)이 원천: 이자부부채 / 리스부채 / 제외 / 현금및현금성자산 / 단기금융상품. "
                "'(기타)금융부채'는 주석 대조 후 이자부 성격이면 F열을 '이자부부채'로 바꾸면 소계·피어·WACC 시트가 다시 계산된다. 리스부채 포함 여부는 요약!B7.")
    ws["B3"].font = note
    for col, title in zip("BCDEFG", ["회사명", "기준시점", "계정과목", "금액", "분류", "비고"]):
        c = ws[f"{col}4"]; c.value = title; c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill, c.border, c.alignment = hdr_fill, box, Alignment(horizontal="center", vertical="center")
    dv = DataValidation(type="list", formula1='"' + ",".join(_CLASSES) + '"', allow_blank=False)
    ws.add_data_validation(dv)
    total_rows = {}
    r = 5
    for p in peers:
        rows = []
        for it in p.get("liab_items", []):
            inc = bool(it.get("include", it.get("default")))
            bucket = next(k for k, v in _pl._bucket_debt([(it["name"], 1)]).items() if v == 1)
            cls = ("리스부채" if bucket == "리스부채" else "이자부부채") if inc else "제외"
            rows.append((it["name"], it["amount"], cls, bucket))
        for it in p.get("cash_items", []):
            rows.append((it["name"], it["amount"], it["kind"], ""))
        if not rows:
            rows.append(("(재무상태표 계정 없음)", 0, "제외", ""))
        a = r
        for k, (nm, amt, cls, bucket) in enumerate(rows):
            if k == 0:
                w.val(ws, r, 2, p["name"]).font = F10B
                w.val(ws, r, 3, fs_date)
                w.val(ws, r, 7, f"{p.get('report_label', '')} · rcept {p.get('rcept_no') or '-'} · {p.get('fs_div') or ''}").font = link
            w.val(ws, r, 4, nm); w.val(ws, r, 5, round(amt / 1e6, 1), fmt=AMT)
            w.val(ws, r, 6, cls, inp=True); dv.add(f"F{r}")
            w.val(ws, r, 8, p["code"]); w.val(ws, r, 9, bucket)
            for col in range(2, 8):
                c = ws.cell(r, col); c.border = box
                if col not in (2, 7):
                    c.font = F10
            r += 1
        b = r - 1
        subs = [("이자부부채 합계(리스 제외)", f'=SUMIFS(E{a}:E{b},F{a}:F{b},"이자부부채")'),
                ("리스부채 합계", f'=SUMIFS(E{a}:E{b},F{a}:F{b},"리스부채")'),
                ("이자부부채 합계", None),
                ("현금및현금성자산 합계", f'=SUMIFS(E{a}:E{b},F{a}:F{b},"현금및현금성자산")'),
                ("단기금융상품 합계", f'=SUMIFS(E{a}:E{b},F{a}:F{b},"단기금융상품")'),
                ("순차입금", None),
                ("비지배지분", None)]
        t = {}
        for label, formula in subs:
            t[label] = r
            for col in range(2, 8):
                c = ws.cell(r, col); c.fill, c.border, c.font = sub_fill, box, F10
            w.val(ws, r, 4, label).font = F10B
            if label == "이자부부채 합계":
                formula = f'=E{t["이자부부채 합계(리스 제외)"]}+IF(요약!$B$7=TRUE,E{t["리스부채 합계"]},0)'
            elif label == "순차입금":
                formula = f'=E{t["이자부부채 합계"]}-E{t["현금및현금성자산 합계"]}-E{t["단기금융상품 합계"]}'
            if formula:
                w.f(ws, r, 5, formula, AMT).font = Font(name=FONT, size=10, bold=True, color="203864")
            else:
                w.val(ws, r, 5, round((p.get("nci") or 0) / 1e6, 1), fmt=AMT).font = F10B
            w.val(ws, r, 6, "소계"); w.val(ws, r, 8, p["code"])
            for col in range(2, 8):
                ws.cell(r, col).fill, ws.cell(r, col).border = sub_fill, box
            r += 1
        total_rows[p["code"]] = t["이자부부채 합계"]
        r += 1
    ws.freeze_panes = "B5"
    return total_rows


KOFIA_URL = "https://www.kofiabond.or.kr"


def _kofia_sheet(w, d):
    """금투협 원자료 시트 — 받아온 표(최종호가수익률·시가평가 매트릭스)를 그대로 싣고,
    WACC 에 쓴 셀만 노란색(FFFF00). 반환 {"rf": "B7", "kd": "M14"} 셀 주소(원자료 없으면 None)."""
    rates = d.get("rates") or {}
    final, matrix = rates.get("final_quotes"), rates.get("matrix")
    if not (final or matrix):
        return None
    ws = w.sheet("금투협")
    used_fill = PatternFill("solid", fgColor="FFFF00")
    ws["A1"] = f"금투협 채권정보센터 조회 원자료 — 적용일 {rates.get('date_used', '')}"
    ws["A1"].font = Font(name=FONT, size=13, bold=True)
    ws["A2"] = "www.kofiabond.or.kr — 최종호가수익률 · 시가평가 기준수익률(평가사 평균)"
    ws["A2"].hyperlink = KOFIA_URL
    ws["A2"].font = Font(name=FONT, size=10, color="0563C1", underline="single")
    ws["A3"] = "노란 셀 = WACC 시트 사용 값 (Rf: 최종호가 국고채권(10년) · Kd: 시가평가 회사채Ⅰ 무보증 BBB- 5년). 단위 %."
    ws["A3"].font = Font(name=FONT, size=9, color="595959")
    refs, r = {}, 5
    if final:
        ws.cell(r, 1, "최종호가수익률 (%)").font = BOLD
        w.header(ws, r + 1, ["종목명", "수익률(%)"])
        r += 2
        for name, v in final.items():
            w.val(ws, r, 1, name)
            c = w.val(ws, r, 2, v, fmt="0.000")
            if name == "국고채권(10년)":
                c.fill = used_fill
                refs["rf"] = f"B{r}"
            r += 1
    if matrix:
        r += 1
        ws.cell(r, 1, "시가평가 기준수익률 매트릭스 — 평가사 평균 (%)").font = BOLD
        w.header(ws, r + 1, ["구분", "종류", "등급"] + KOFIA_MATURITIES)
        r += 2
        for row in matrix:
            for col, key in enumerate(["category", "type", "grade"], 1):
                w.val(ws, r, col, row.get(key, ""))
            for j, mat in enumerate(KOFIA_MATURITIES, 4):
                c = w.val(ws, r, j, row.get(mat), fmt="0.000")
                if (str(row.get("category", "")).startswith("회사채 I") and row.get("type") == "무보증"
                        and row.get("grade") == "BBB-" and mat == "5년"):
                    c.fill = used_fill
                    refs["kd"] = f"{get_column_letter(j)}{r}"
            r += 1
    w.widths(ws, [24, 12, 12] + [8] * len(KOFIA_MATURITIES))
    ws.freeze_panes = "A5"
    return refs


def _target_de(w, peers):
    ws = w.sheet("목표DE")
    w.header(ws, 1, ["회사명", "종목코드", "βU(집계분)", "D/E(집계분)"])
    for r, p in enumerate(peers, 2):
        w.val(ws, r, 1, p["name"]); w.val(ws, r, 2, p["code"])
        w.f(ws, r, 3, f'=IF(피어!Y{r}=TRUE,피어!X{r},"")', "0.0000")
        w.f(ws, r, 4, f'=IF(피어!Y{r}=TRUE,피어!S{r},"")', "0.0000")
    w.widths(ws, [18, 9, 12, 12])


def _excel_recalc(path):
    """openpyxl 파일엔 계산값이 없어 수동 계산 모드 Excel 에선 빈칸으로 보인다.
    Excel(COM)로 열어 전체 계산 후 저장해 값을 심는다. Excel 이 없으면 조용히 건너뛴다."""
    try:
        import win32com.client
        xl = win32com.client.DispatchEx("Excel.Application")
    except Exception:
        return False
    try:
        xl.Visible = False; xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(path, UpdateLinks=0)
        xl.CalculateFull()
        wb.Save(); wb.Close(False)
        return True
    except Exception:
        return False
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


def _save_with_fallback(wb, path):
    base, ext = os.path.splitext(path)
    cand, n = path, 1
    while True:
        try:
            wb.save(cand)
            return cand
        except PermissionError:
            n += 1
            cand = f"{base}_v{n}{ext}"
            if n > 20:
                raise


def build(data, path, recalc=False):
    """recalc=True 면 저장 후 Excel 로 열어 계산값을 심는다(앱에서 사용). 테스트는 False."""
    w = _W()
    peers = list(data.get("peers", []))
    _summary(w, data)
    kofia_refs = _kofia_sheet(w, data)
    _wacc(w, data, peers, kofia_refs=kofia_refs)
    _candidates(w, data)
    total_rows = _debt_schedule(w, data, peers)
    _peers(w, data, peers, total_rows)
    ranges = _prices(w, data, peers)
    _beta(w, data, peers, ranges)
    _target_de(w, peers)
    # 시트 순서 고정: 요약·후보군·피어·주가·베타·목표DE (·금투협 원자료)
    order = ["요약", "WACC", "후보군", "피어", DEBT_SHEET, "주가", "베타", "목표DE"] + (["금투협"] if kofia_refs is not None else [])
    w.wb._sheets = [w.wb[n] for n in order]
    w.wb.calculation.fullCalcOnLoad = True      # openpyxl 파일엔 계산값이 없으므로 열 때 전체 재계산을 강제
    saved = _save_with_fallback(w.wb, path)
    if recalc:
        _excel_recalc(saved)
    return saved, w.formulas
