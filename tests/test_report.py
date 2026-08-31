import os, tempfile, unittest

import openpyxl

import report


def _data():
    dates = [f"2026-01-{d:02d}" for d in (2, 9, 16, 23, 30)]
    return {
        "target": {"name": "테스트대상", "code": "", "industry": "밸브 제조업", "ksic": "291"},
        "as_of": "2026-01-31", "beta_source": "kicpa", "de_method": "mean", "tax_target": 0.275,
        "keywords": ["밸브", "피팅"], "query_date": "2026-08-27", "window": ("20240101", "20260130"),
        "candidates": [
            {"name": "성광벤드", "market": "유가", "code": "014620", "industry": "기타 금속 가공제품 제조업",
             "products": "관이음쇠(피팅)", "listed": "1997-11-14", "settle_month": "12월",
             "hits": ["피팅"], "flags": [], "excluded": False, "reason": ""},
            {"name": "스팩A", "market": "코스닥", "code": "400000", "industry": "금융 지원 서비스업",
             "products": "밸브 인수합병", "listed": "2025-01-01", "settle_month": "12월",
             "hits": ["밸브"], "flags": ["스팩", "상장2년미만"], "excluded": True, "reason": "스팩"},
        ],
        "peers": [
            {"name": "성광벤드", "code": "014620", "ksic": "25919", "market": "유가",
             "close_krx": 10000.0, "close_kicpa": 10050.0, "common_out": 1000, "pref_out": 0, "nci": 0,
             "debt": {"단기차입금": 100, "유동성장기부채": 0, "장기차입금": 200, "사채": 0, "리스부채": 50, "기타": 0},
             "tax": 0.275, "include": True, "flags": [], "rcept_no": "2026031700001", "report_label": "2025 사업보고서", "fs_div": "CFS",
             "liab_items": [{"name": "단기차입금", "amount": 100, "default": True, "include": True},
                            {"name": "장기차입금", "amount": 200, "default": True, "include": True},
                            {"name": "리스부채", "amount": 50, "default": True, "include": True},
                            {"name": "유동금융부채", "amount": 70, "default": False, "include": False}],
             "cash_items": [{"name": "현금및현금성자산", "amount": 30, "kind": "현금및현금성자산"}],
             "kicpa": {"base_date": "2026-01-30", "close": 10050.0, "raw": 0.6, "adjusted": 0.73, "points": 104}},
            {"name": "신규상장", "code": "0155E0", "ksic": "", "market": "코스닥",
             "close_krx": None, "close_kicpa": None, "common_out": 0, "pref_out": 0, "nci": 0,
             "debt": {"단기차입금": 0, "유동성장기부채": 0, "장기차입금": 0, "사채": 0, "리스부채": 0, "기타": 0},
             "tax": 0.275, "include": False, "flags": ["임시코드", "데이터 없음"], "rcept_no": None, "report_label": "", "fs_div": None,
             "kicpa": None},
        ],
        "weekly_dates": dates, "weekly_index": [2500.0, 2520.0, 2490.0, 2530.0, 2550.0],
        "weekly": {"014620": [9800.0, 9900.0, 9700.0, 10000.0, 10000.0], "0155E0": [None, None, None, 5000.0, 5100.0]},
        "notes": ["주가: KRX 수정종가, 배당락 미보정"],
    }


def _raw_rates():
    """kofia.market_rates 결과(요약 + 원자료). 매트릭스 만기 키는 일부만 채운다."""
    mat = [
        {"category": "국채", "type": "국고채권", "grade": "양곡,외평,재정", "3년": 3.555, "5년": 3.771, "10년": 3.877},
        {"category": "회사채 I(공모사채)", "type": "무보증", "grade": "BBB-", "3년": 9.961, "5년": 10.46, "10년": 11.3},
        {"category": "회사채 I(공모사채)", "type": "무보증", "grade": "AA-", "3년": 4.166, "5년": 4.4, "10년": 4.8},
    ]
    final = {"국고채권(10년)": 3.879, "회사채(무보증3년)BBB-": 9.961, "CD수익률(91일)": 2.82}
    return {"date_used": "2026-03-31", "ktb10_final": 3.879, "ktb10_val": 3.877,
            "bbb_minus_5y": 10.46, "bbb_minus_3y": 9.961, "aa_minus_5y": 4.4,
            "final_quotes": final, "matrix": mat}


class KofiaSheetTest(unittest.TestCase):
    """금투협 원자료 시트: 받아온 표 그대로 + 사용 셀 노란색 + 하이퍼링크. WACC Rf·Kd는 이 시트를 참조."""

    def setUp(self):
        d = _data(); d["rates"] = _raw_rates()
        self.path, _ = report.build(d, os.path.join(tempfile.mkdtemp(), "k.xlsx"))
        self.wb = openpyxl.load_workbook(self.path)

    def test_sheet_exists_last_with_link(self):
        self.assertEqual(self.wb.sheetnames[-1], "금투협")
        ws = self.wb["금투협"]
        self.assertIn("금투협", str(ws["A1"].value))
        self.assertTrue(str(ws["A2"].hyperlink.target).startswith("https://www.kofiabond.or.kr"))

    def test_final_quotes_table_and_used_cell_yellow(self):
        ws = self.wb["금투협"]
        # 최종호가: 헤더 6행, 자료 7행~ (국고채권(10년)이 첫 행)
        self.assertEqual(ws["A6"].value, "종목명")
        self.assertEqual(ws["A7"].value, "국고채권(10년)"); self.assertEqual(ws["B7"].value, 3.879)
        self.assertEqual(ws["B7"].fill.fgColor.rgb, "00FFFF00")
        self.assertEqual(ws["A9"].value, "CD수익률(91일)"); self.assertNotEqual(ws["B9"].fill.fgColor.rgb, "00FFFF00")

    def test_matrix_table_and_used_cell_yellow(self):
        ws = self.wb["금투협"]
        # 매트릭스: 섹션 11행, 헤더 12행(구분·종류·등급 + 16개 만기), 자료 13~15행
        self.assertEqual([ws.cell(12, c).value for c in range(1, 4)], ["구분", "종류", "등급"])
        self.assertEqual(ws.cell(12, 13).value, "5년")
        self.assertEqual(ws["C14"].value, "BBB-"); self.assertEqual(ws["M14"].value, 10.46)
        self.assertEqual(ws["M14"].fill.fgColor.rgb, "00FFFF00")
        self.assertEqual(ws["M13"].value, 3.771)                     # 국고 5년은 사용 안 함 → 노랑 아님
        self.assertNotEqual(ws["M13"].fill.fgColor.rgb, "00FFFF00")

    def test_wacc_refers_to_kofia_sheet(self):
        ws = self.wb["WACC"]
        self.assertEqual(ws["C16"].value, "=금투협!B7/100")
        self.assertEqual(ws["C20"].value, "=금투협!M14/100")
        self.assertEqual(ws["C16"].fill.fgColor.rgb, "00FFF2CC")     # 여전히 입력 스타일(수정 가능)
        self.assertIn("금투협", ws["D16"].value); self.assertIn("BBB- 5년", ws["D20"].value)


class SrpAutoTest(unittest.TestCase):
    """WACC 시트 SRP: 대상 시총(상장)·순자산(비상장)으로 한공회 5분위 자동 판정."""

    def _build(self, **target_extra):
        d = _data(); d["target"].update(target_extra)
        p, _ = report.build(d, os.path.join(tempfile.mkdtemp(), "s.xlsx"))
        return openpyxl.load_workbook(p)["WACC"]

    def test_cap_quintile_1(self):
        ws = self._build(cap_eok=20_000)                             # 2조 → 1분위 -0.51%
        self.assertAlmostEqual(ws["C18"].value, -0.0051)
        self.assertIn("1분위", ws["D18"].value); self.assertIn("시가총액", ws["D18"].value)

    def test_net_assets_quintile_4(self):
        ws = self._build(net_assets_eok=2_300)                       # 순자산 2,300억 → 4분위 2.67%
        self.assertAlmostEqual(ws["C18"].value, 0.0267)
        self.assertIn("4분위", ws["D18"].value); self.assertIn("순자산", ws["D18"].value)

    def test_default_stays_486(self):
        ws = self._build()
        self.assertAlmostEqual(ws["C18"].value, 0.0486)
        self.assertIn("5분위", ws["D18"].value)


class FsAsOfSheetTest(unittest.TestCase):
    def test_debt_schedule_shows_fs_as_of(self):
        d = _data(); d["fs_as_of"] = "2025-12-31"
        p, _ = report.build(d, os.path.join(tempfile.mkdtemp(), "f.xlsx"))
        dd = openpyxl.load_workbook(p)["이자부부채 산정내역"]
        self.assertEqual(dd["C5"].value, "2025-12-31")
        self.assertIn("재무 기준일 2025-12-31", dd["B2"].value)


class ReportTest(unittest.TestCase):
    def setUp(self):
        self.path, self.formulas = report.build(_data(), os.path.join(tempfile.mkdtemp(), "out.xlsx"))
        self.wb = openpyxl.load_workbook(self.path)

    def test_sheet_order(self):
        self.assertEqual(self.wb.sheetnames, ["요약", "WACC", "후보군", "피어", "이자부부채 산정내역", "주가", "베타", "목표DE"])
        self.assertEqual(self.wb["요약"]["A7"].value, "리스부채 포함(D)"); self.assertEqual(self.wb["요약"]["B7"].value, True)

    def test_summary_cells_are_formulas_and_inputs_are_values(self):
        ws = self.wb["요약"]
        self.assertEqual(ws["B4"].value, "한공회")
        self.assertEqual(ws["B5"].value, "평균")
        self.assertEqual(ws["B6"].value, 0.275)
        self.assertEqual(ws["B10"].value, "=B8*(1+(1-B6)*B9)")
        self.assertTrue(str(ws["B8"].value).startswith("=IF($B$5="))
        self.assertEqual(ws["B4"].fill.fgColor.rgb, "00FFF2CC")

    def test_candidates_keyword_formulas(self):
        ws = self.wb["후보군"]
        self.assertEqual([ws.cell(1, c).value for c in range(1, 8)],
                         ["회사명", "시장", "종목코드", "업종", "주요제품", "밸브", "피팅"])
        self.assertEqual(ws["F2"].value, '=IFERROR(FIND(F$1,$E2&" "&$D2),0)>0')
        self.assertEqual(ws["H1"].value, "제외")
        self.assertEqual(ws["H3"].value, True)
        self.assertEqual(ws["I3"].value, "스팩")
        self.assertEqual(ws["J2"].value, "=AND(OR(F2:G2),NOT(H2=TRUE))")

    def test_peer_row_formulas(self):
        ws = self.wb["피어"]
        self.assertEqual(ws["B2"].value, "014620")
        self.assertEqual(ws["G2"].value, '=IF(AND(요약!$B$4="한공회",F2<>""),F2,E2)')
        self.assertEqual(ws["K2"].value, "=G2*(H2+I2)+J2")
        self.assertEqual(ws["L2"].value, "=SUMIFS('이자부부채 산정내역'!$E:$E,'이자부부채 산정내역'!$H:$H,$B2,'이자부부채 산정내역'!$I:$I,\"단기차입금\",'이자부부채 산정내역'!$F:$F,\"이자부부채\")*1000000")
        self.assertEqual(ws["P2"].value, "=SUMIFS('이자부부채 산정내역'!$E:$E,'이자부부채 산정내역'!$H:$H,$B2,'이자부부채 산정내역'!$I:$I,\"리스부채\",'이자부부채 산정내역'!$F:$F,\"리스부채\")*1000000")
        dd = self.wb["이자부부채 산정내역"]
        self.assertEqual([dd.cell(4, c).value for c in range(2, 8)], ["회사명", "기준시점", "계정과목", "금액", "분류", "비고"])
        # 성광벤드 블록: 5~9행 계정(4 부채 + 1 현금), 10~16행 소계
        self.assertEqual(dd["B5"].value, "성광벤드"); self.assertEqual(dd["C5"].value, "2026-01-31")
        self.assertEqual([dd.cell(r, 4).value for r in range(5, 10)], ["단기차입금", "장기차입금", "리스부채", "유동금융부채", "현금및현금성자산"])
        self.assertEqual([dd.cell(r, 6).value for r in range(5, 10)], ["이자부부채", "이자부부채", "리스부채", "제외", "현금및현금성자산"])
        self.assertEqual(dd["F8"].fill.fgColor.rgb, "00FFF2CC")
        self.assertEqual(dd["H5"].value, "014620"); self.assertEqual(dd["I5"].value, "단기차입금")
        self.assertEqual(dd["D10"].value, "이자부부채 합계(리스 제외)"); self.assertEqual(dd["E10"].value, '=SUMIFS(E5:E9,F5:F9,"이자부부채")')
        self.assertEqual(dd["D12"].value, "이자부부채 합계"); self.assertEqual(dd["E12"].value, "=E10+IF(요약!$B$7=TRUE,E11,0)")
        self.assertEqual(dd["D15"].value, "순차입금"); self.assertEqual(dd["E15"].value, "=E12-E13-E14")
        self.assertEqual(ws["R2"].value, "='이자부부채 산정내역'!E12*1000000")
        self.assertEqual(ws["X2"].value, '=IF(OR(W2="",S2=""),"",W2/(1+(1-T2)*S2))')
        self.assertEqual(ws["Y2"].value, True)
        self.assertEqual(ws["Y3"].value, False)
        self.assertIsNone(ws["E3"].value)

    def test_price_and_beta_formulas(self):
        px, bt = self.wb["주가"], self.wb["베타"]
        self.assertEqual(px["A1"].value, "주차 종료일")
        self.assertEqual(px["C1"].value, "성광벤드")
        self.assertEqual(px["B2"].value, 2500.0)
        self.assertEqual(px["F1"].value, "KOSPI 수익률")           # 빈 열 E 다음
        self.assertEqual(px["F3"].value, '=IF(OR(B2="",B3=""),"",B3/B2-1)')
        self.assertEqual(px["G3"].value, '=IF(OR(C2="",C3=""),"",C3/C2-1)')
        self.assertEqual(bt["C2"].value, "=SLOPE(주가!$G$3:$G$6,주가!$F$3:$F$6)")
        self.assertEqual(bt["E2"].value, "=COUNT(주가!$G$3:$G$6)")
        self.assertEqual(bt["F2"].value, '=IF(C2="","",2/3*C2+1/3)')
        self.assertEqual(bt["K2"].value, 0.73)
        self.assertIsNone(bt["K3"].value)

    def test_target_de_sheet(self):
        ws = self.wb["목표DE"]
        self.assertEqual(ws["C2"].value, '=IF(피어!Y2=TRUE,피어!X2,"")')
        self.assertEqual(ws["D2"].value, '=IF(피어!Y2=TRUE,피어!S2,"")')

    def test_wacc_sheet_layout(self):
        ws = self.wb["WACC"]
        self.assertTrue(str(ws["A1"].value).strip().startswith("WACC 산출 — 테스트대상"))
        self.assertEqual(ws["A1"].fill.fgColor.rgb, "001F3864"); self.assertEqual(ws["A1"].font.color.rgb, "00FFFFFF")
        self.assertFalse(ws.sheet_view.showGridLines)
        # ① 피어 표: 헤더 5행, 피어 6~7행(2개), 평균 8행
        self.assertEqual(ws["B5"].value, "회사명"); self.assertEqual(ws["B5"].fill.fgColor.rgb, "00C6D9F1")
        self.assertEqual(ws["B6"].value, "성광벤드"); self.assertEqual(ws["C6"].value, "014620")
        self.assertEqual(ws["D6"].value, '=IF(피어!Y2=TRUE,"O","")')
        self.assertEqual(ws["E6"].value, "=피어!W2"); self.assertEqual(ws["F6"].value, "=피어!K2/1000000"); self.assertEqual(ws["H6"].value, "=피어!R2/1000000")
        self.assertEqual(ws["J6"].value, '=IF(D6="O",피어!S2,"")'); self.assertEqual(ws["L6"].value, "=피어!T2")
        self.assertEqual(ws["N6"].value, '=IFERROR(E6/(1+(1-L6)*J6),"")'); self.assertTrue(ws["N6"].font.bold)
        self.assertEqual(ws["B8"].value, "평균 (집계 대상 O)")
        self.assertEqual(ws["N8"].value, '=IF(요약!$B$5="중앙값",MEDIAN(N6:N7),AVERAGE(N6:N7))')
        self.assertEqual(ws["J8"].value, '=IF(요약!$B$5="중앙값",MEDIAN(J6:J7),AVERAGE(J6:J7))')
        # ② 자본비용: 섹션 10행, 헤더 11행, 항목 12~24행
        self.assertEqual(ws["B10"].value, "②  자본비용 산출 (Cost of Capital)")
        self.assertEqual(ws["B12"].value, "언레버 베타 (βU, 피어 평균)"); self.assertEqual(ws["C12"].value, "=N8")
        self.assertEqual(ws["C13"].value, "=J8"); self.assertEqual(ws["C14"].value, "=요약!$B$6")
        self.assertEqual(ws["C15"].value, "=C12*(1+(1-C14)*C13)")
        self.assertEqual(ws["B16"].value, "무위험이자율 (Rf)"); self.assertEqual(ws["C16"].fill.fgColor.rgb, "00FFF2CC")
        self.assertEqual(ws["C17"].value, 0.08); self.assertEqual(ws["C18"].value, 0.0486)
        self.assertEqual(ws["C19"].value, "=C16+C15*C17+C18")
        self.assertEqual(ws["C20"].fill.fgColor.rgb, "00FFF2CC")                       # Kd 입력
        self.assertEqual(ws["C21"].value, "=C20*(1-C14)")
        self.assertEqual(ws["C22"].value, "=1/(1+C13)"); self.assertEqual(ws["C23"].value, "=C13/(1+C13)")
        self.assertEqual(ws["B24"].value, "WACC"); self.assertEqual(ws["C24"].value, "=ROUND(C19*C22+C21*C23,4)")
        self.assertEqual(ws["C24"].fill.fgColor.rgb, "00C6D9F1")
        # 수식 스트립과 콜아웃
        self.assertEqual(ws["F11"].value, "Rf"); self.assertEqual(ws["N12"].value, "=C19")
        self.assertEqual(ws["F15"].value, "Ke"); self.assertEqual(ws["N16"].value, "=C24")
        self.assertEqual(ws["K19"].value, "=C24"); self.assertEqual(ws["K19"].font.size, 16)
        self.assertEqual(ws["B6"].font.name, "맑은 고딕"); self.assertEqual(ws["B6"].font.size, 9)

    def test_wacc_rates_prefill(self):
        d = _data(); d["rates"] = {"date_used": "2026-03-31", "ktb10_final": 3.879, "ktb10_val": 3.877, "bbb_minus_5y": 10.46, "bbb_minus_3y": 9.961, "aa_minus_5y": 4.4}
        p, _ = report.build(d, os.path.join(tempfile.mkdtemp(), "r.xlsx"))
        ws = openpyxl.load_workbook(p)["WACC"]
        self.assertAlmostEqual(ws["C16"].value, 0.03879); self.assertIn("최종호가수익률 국고채 10년 3.879%", ws["D16"].value)
        self.assertAlmostEqual(ws["C20"].value, 0.1046); self.assertIn("BBB- 5년 10.460%", ws["D20"].value)
        self.assertEqual(ws["C16"].fill.fgColor.rgb, "00FFF2CC")               # 여전히 입력 셀(수정 가능)

    def test_every_equals_string_is_a_tracked_formula(self):
        seen = set()
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.data_type == "f":
                        seen.add(f"{ws.title}!{c.coordinate}")
                    elif isinstance(c.value, str) and c.value.startswith("="):
                        self.assertEqual(c.value, "=", f"{ws.title}!{c.coordinate} 수식 아닌 '='로 시작하는 텍스트")
        self.assertEqual(seen, self.formulas)
        self.assertIn("요약!B10", self.formulas)

    def test_permission_fallback(self):
        d = tempfile.mkdtemp()
        p = os.path.join(d, "x.xlsx")
        with open(p, "w") as f:
            f.write("lock")
        from unittest import mock
        real = openpyxl.Workbook.save
        calls = {"n": 0}
        def flaky(self_, filename):
            calls["n"] += 1
            if filename == p:
                raise PermissionError("locked")
            return real(self_, filename)
        with mock.patch.object(openpyxl.Workbook, "save", flaky):
            saved, _ = report.build(_data(), p)
        self.assertEqual(saved, os.path.join(d, "x_v2.xlsx"))
